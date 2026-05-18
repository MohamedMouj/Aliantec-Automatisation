import openpyxl
from openpyxl.styles import PatternFill
import re

class ExcelHelper:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.data=None

    def write_data_to_excel(self, data):
        wb = openpyxl.Workbook()
        ws_E = wb.create_sheet("E")
        ws_IC = wb.create_sheet("IC")
        ws_others = wb.create_sheet("Others")
        ws_E.append(["Fuseaux","File Name", "NV", "CC", "APPAREIL"])
        ws_IC.append(["Fuseaux","File Name", "NV", "CC", "APPAREIL"])
        ws_others.append(["Fuseaux","File Name", "NV", "CC", "APPAREIL"])
        ws_distinct = wb.create_sheet("DISTINCT (NO DOUBLONS)")
        ws_distinct.append(["NV", "CC", "APPAREIL"])


        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")        
        yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") 
        
        prev_file = None
        color_cycle = [red_fill, yellow_fill, green_fill]
        color_index = -1
        device_set = set()
        
        for row in data:
            fuseaux = row[0]
            filename = row[1]
            sheet_name= row[2]
            tokens = row[3:]
            
            if filename != prev_file:
                color_index = (color_index + 1) % len(color_cycle)
                prev_file = filename
            
            current_fill = color_cycle[color_index]
            
            if isinstance(tokens, list):
                if "IC"==sheet_name:
                    ws_IC.append([fuseaux]+[filename] + tokens)
                    current_row_idx = ws_IC.max_row
                    for cell in ws_IC[current_row_idx]:
                        cell.fill = current_fill
                elif "E"==sheet_name:
                    ws_E.append([fuseaux]+[filename] + tokens)
                    current_row_idx = ws_E.max_row
                    for cell in ws_E[current_row_idx]:
                        cell.fill = current_fill
                elif "others"==sheet_name:
                    ws_others.append([fuseaux]+[filename] + tokens)
                    current_row_idx = ws_others.max_row
                    for cell in ws_others[current_row_idx]:
                        cell.fill = current_fill
                    
                    
                    val = str(tokens[-1])
                    if not (val.startswith('B') or val[:3] in ["VSM", "UDB", "UFM"]):
                        device_set.add((str(tokens[0]), str(tokens[1]), val))
        

        
        list_distinct=list(device_set)


             

       
        sorted_devices = sorted(list_distinct, key=lambda x: x[2])

        for device in sorted_devices:
            ws_distinct.append([device[0], device[1], device[2]])

        wb.save(self.excel_file)
    
        