import os
import re
import tempfile
import zipfile

import openpyxl
import pandas as pd


class ExcelParser:
    MAIN_SHEET = "Validation Web"
    EXTRACT_SHEET = "Extraction VIN"

    def __init__(self, excel_file_name):
        self.excel_file_name = excel_file_name

        # Main sheet: loaded normally (not read-only) so its existing layout -
        # styles, merged cells, column widths, formulas, etc. - is preserved.
        # We edit its cells in place rather than rebuilding it from raw values.
        self.wb = None
        self.main_sheet = None

        # Extraction sheet: loaded via pandas only (header=None so col/row
        # indices are 0-based and map directly to excel col/row - 1). It is
        # never parsed into openpyxl's cell/style object model.
        self.extract_df = None
        self.header_row = None

        # Memoized {excel_col_number: {stripped_value: first_excel_row}} index,
        # built lazily and only for columns that are actually queried.
        self._col_index_cache = {}

        # Path to a temp copy of the workbook with EXTRACT_SHEET's XML
        # emptied out, so openpyxl never has to parse it. Cleaned up in close().
        self._stripped_path = None

    # ------------------------------------------------------------------ #
    # Zip-level sheet stripping - openpyxl has no native "load only sheet X"
    # like pandas' sheet_name=, since normal mode always builds full Cell
    # objects for every sheet in the file. This works around that by editing
    # the .xlsx archive itself before openpyxl ever sees it.
    # ------------------------------------------------------------------ #
    def _find_sheet_xml_path(self, sheet_name):
        with zipfile.ZipFile(self.excel_file_name) as z:
            workbook_xml = z.read("xl/workbook.xml").decode("utf-8")
            rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        sheet_tag_match = re.search(
            rf'<sheet\b[^>]*name="{re.escape(sheet_name)}"[^>]*/?>', workbook_xml
        )
        if not sheet_tag_match:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook.xml")
        sheet_tag = sheet_tag_match.group(0)

        rid_match = re.search(r'r:id="(rId\d+)"', sheet_tag)
        if not rid_match:
            raise KeyError(f"No r:id found for sheet '{sheet_name}' in workbook.xml")
        rid = rid_match.group(1)

        rel_tag_match = re.search(
            rf'<Relationship\b[^>]*Id="{rid}"[^>]*/?>', rels_xml
        )
        if not rel_tag_match:
            raise KeyError(f"Relationship '{rid}' not found in workbook.xml.rels")
        rel_tag = rel_tag_match.group(0)

        target_match = re.search(r'Target="([^"]+)"', rel_tag)
        if not target_match:
            raise KeyError(f"No Target found for relationship '{rid}'")
        target = target_match.group(1).lstrip("/")

        return target if target.startswith("xl/") else f"xl/{target}"

    def _strip_sheet_to_temp(self, sheet_name):
        sheet_path = self._find_sheet_xml_path(sheet_name)
        empty_sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            "<sheetData/></worksheet>"
        )

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)

        with zipfile.ZipFile(self.excel_file_name, "r") as zin, zipfile.ZipFile(
            tmp_path, "w", zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == sheet_path:
                    # Skip reading/decompressing the huge original entry
                    # entirely - just write the tiny placeholder instead.
                    zout.writestr(item, empty_sheet_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        return tmp_path

    # ------------------------------------------------------------------ #
    # Loading
    # ------------------------------------------------------------------ #
    def load_excel(self):
        # openpyxl loads this stripped copy, where EXTRACT_SHEET is an
        # empty placeholder - so it only ever parses/builds Cell objects
        # for MAIN_SHEET (with all of its original formatting intact).
        self._stripped_path = self._strip_sheet_to_temp(self.EXTRACT_SHEET)
        self.wb = openpyxl.load_workbook(self._stripped_path, data_only=True)

        if self.MAIN_SHEET not in self.wb.sheetnames:
            raise KeyError(
                f"Sheet '{self.MAIN_SHEET}' not found. Available: {self.wb.sheetnames}"
            )
        if self.EXTRACT_SHEET not in self.wb.sheetnames:
            raise KeyError(
                f"Sheet '{self.EXTRACT_SHEET}' not found. Available: {self.wb.sheetnames}"
            )

        self.main_sheet = self.wb[self.MAIN_SHEET]

        # The placeholder EXTRACT_SHEET is empty and cheap to drop now -
        # it never needs to exist in the workbook object at all.
        del self.wb[self.EXTRACT_SHEET]

        # Fast, vectorized lookups for the extraction sheet - read straight
        # from the ORIGINAL, untouched file via pandas only.
        self.extract_df = pd.read_excel(
            self.excel_file_name,
            sheet_name=self.EXTRACT_SHEET,
            header=None,
            dtype=str,
        )
        self.header_row = self.extract_df.iloc[0]

    def close(self):
        if self.wb:
            self.wb.close()
        if self._stripped_path and os.path.exists(self._stripped_path):
            os.remove(self._stripped_path)

    # ------------------------------------------------------------------ #
    # Label parsing (unchanged - pure string logic, not a bottleneck)
    # ------------------------------------------------------------------ #
    def extract_label(self, cell_value):
        if cell_value is None:
            return None
        val = str(cell_value).strip()
        if val == "":
            return None

        matched = re.findall(r"\b(?=[A-Z]*[0-9])[A-Z0-9]{5}\b", val)
        if matched:
            return matched
        elif "Générique" in val:
            return ["DXD00"]
        return None

    # ------------------------------------------------------------------ #
    # Column search on the extraction sheet's header row
    # ------------------------------------------------------------------ #
    def search_col_by_label(self, label):
        cols = []
        for lab in label:
            prefix = lab[:3]
            for idx, val in self.header_row.items():
                if pd.notna(val) and str(val).strip().startswith(prefix):
                    cols.append(idx + 1)  # 1-indexed, like openpyxl columns
        return cols if cols else None

    # ------------------------------------------------------------------ #
    # Fast per-column index (built once per column, cached)
    # ------------------------------------------------------------------ #
    def _get_col_index(self, col):
        """
        Return {stripped_value: first_excel_row} for the given 1-indexed
        excel column, built with vectorized pandas ops and memoized so a
        column is only ever scanned once, no matter how many main-sheet
        rows reference it.
        """
        if col in self._col_index_cache:
            return self._col_index_cache[col]

        df_col = col - 1
        s = self.extract_df.iloc[:, df_col]
        s = s.dropna()
        s = s.astype(str).str.strip()
        s = s[s != ""]
        s = s[~s.duplicated(keep="first")]  # keep first (topmost) occurrence

        value_to_row = dict(zip(s.values, (s.index + 1).tolist()))
        self._col_index_cache[col] = value_to_row
        return value_to_row

    # ------------------------------------------------------------------ #
    # VIN row lookup
    # ------------------------------------------------------------------ #
    def find_vin_row(self, cols, label):
        """
        NOTE on the original logic: it built a python dict/list match across
        every column in `cols`, but its final filter step only ever kept
        matches found in `cols[-1]` (the last matched column) - matches from
        any other column never affected the returned row. This version
        preserves that exact behavior, but skips indexing the other columns
        entirely since they never contributed to the result, which is
        also a meaningful speedup on its own.
        """
        if not cols:
            return None

        last_col = cols[-1]
        map_label_col = {col: lab[3:] for col, lab in zip(cols, label)}
        suffix = map_label_col.get(last_col)
        if suffix is None:
            return None

        if len(suffix) > 1 and suffix.startswith("0"):
            suffix = suffix[1:]
            if len(suffix) > 1 and suffix.startswith("0"):
                suffix = suffix[1:]

        return self._get_col_index(last_col).get(suffix)

    def find_vin(self, row_idx):
        if row_idx is None:
            return None
        df_row = row_idx - 1
        if df_row < 0 or df_row >= len(self.extract_df):
            return None
        val = self.extract_df.iat[df_row, 3]  # column D
        if pd.isna(val):
            return None
        val = str(val).strip()
        return val if val else None

    # ------------------------------------------------------------------ #
    # ELEC / IMP aggregation from the main sheet
    # ------------------------------------------------------------------ #
    def get_data(self):
        elec = {}
        imp = {}
        for row in self.main_sheet.iter_rows(min_row=3):
            for cell in row[8:]:
                if cell.value is not None:
                    sommaire = str(row[0].value)
                    schema = str(row[1].value)
                    if cell.column % 2 != 0:
                        elec[str(cell.value)] = [schema, sommaire]
                    else:
                        imp[str(cell.value)] = [schema, sommaire]
        return [elec, imp]

    def write_elec_data(self, elec):
        ws = self.wb.create_sheet("ELEC")
        ws.append(["Sommaire", "Schema", "ELEC"])
        for code, data in elec.items():
            ws.append([data[0], data[1], code])

    def write_imp_data(self, imp):
        ws = self.wb.create_sheet("IMP")
        ws.append(["Sommaire", "Schema", "IMP"])
        for code, data in imp.items():
            ws.append([data[0], data[1], code])

    # ------------------------------------------------------------------ #
    # Main driver
    # ------------------------------------------------------------------ #
    def start(self, output_path=None):
        results = []

        # Edit main_sheet's cells directly - same sheet object, same styles,
        # merged cells, column widths, etc. Nothing about its layout changes.
        for row in self.main_sheet.iter_rows():
            row_number = row[0].row

            if row_number <= 2:
                continue

            d_cell = row[3]  # column D
            lab = self.extract_label(d_cell.value)

            if not lab:
                continue

            labels = ", ".join(lab)
            cols = self.search_col_by_label(lab)
            if cols is None:
                results.append({
                    "row": row_number,
                    "label": labels,
                    "vin": None,
                    "status": "warn",
                    "message": f"No column found for label '{lab}'",
                })
                continue

            vin_row = self.find_vin_row(cols, lab)
            vin = self.find_vin(vin_row)

            if vin:
                self.main_sheet.cell(row=row_number, column=7).value = vin
                results.append({
                    "row": row_number,
                    "label": labels,
                    "vin": vin,
                    "status": "ok",
                    "message": "VIN written",
                })

        elec, imp = self.get_data()
        self.write_elec_data(elec)
        self.write_imp_data(imp)

        save_path = output_path or self.excel_file_name
        self.wb.save(save_path)
        return results, save_path