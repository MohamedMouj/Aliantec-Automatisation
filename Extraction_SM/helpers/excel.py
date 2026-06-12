import openpyxl
from openpyxl import cell
from openpyxl.styles import Border, Side
import re


class ExcelParser:
    SHEET_NAME = "Sommaire"
    SHEET_NAME_EXTRACTION = "Extraction SM"
    def __init__(self, excel_file_name):
        self.excel_file_name = excel_file_name
        self.wb = None
        self.sheet = None

    def close(self):
        if self.wb:
            self.wb.close()
    def load_excel(self):
        self.wb = openpyxl.load_workbook(self.excel_file_name, data_only=True)

        if self.SHEET_NAME not in self.wb.sheetnames or self.SHEET_NAME_EXTRACTION not in self.wb.sheetnames:
            raise KeyError(
                f"Sheet '{self.SHEET_NAME}' or '{self.SHEET_NAME_EXTRACTION}' not found. "
                f"Available: {self.wb.sheetnames}"
            )

        self.sheet = self.wb[self.SHEET_NAME]
        self.sheet_extraction = self.wb[self.SHEET_NAME_EXTRACTION]

    def close(self):
        if self.wb:
            self.wb.close()

    def determine_rows_to_remove(self):
        last_col = self.sheet.max_column
        rows_to_delete = []
        for cell in list(self.sheet.columns)[last_col - 1]:
            if cell.row <= 6:
                continue
            print(f"Row {cell.row}: value={repr(cell.value)}")  # ← add this
            if cell.value is None or str(cell.value).strip() == "":
                rows_to_delete.append(cell.row)
        print("Rows to delete:", rows_to_delete)  # ← and this
        return rows_to_delete

    def delete_rows(self, rows_to_delete):
        rows_set = set(rows_to_delete)
        # Unmerge any merged cells that touch rows being deleted
        # openpyxl leaves ghost merge definitions if you delete without unmerging first
        for merge in list(self.sheet.merged_cells.ranges):
            if merge.min_row in rows_set or merge.max_row in rows_set:
                self.sheet.unmerge_cells(str(merge))
        # Now safely delete the rows
        for row in sorted(rows_to_delete, reverse=True):
            self.sheet.delete_rows(row)

    def delete_col(self):
        self.sheet.delete_cols(idx=3, amount=7)
    
    def create_col_before_last_column(self):
        last_col_index = self.sheet.max_column  
        self.sheet.insert_cols(last_col_index)
        return list(self.sheet.iter_cols(min_col=last_col_index, max_col=last_col_index))[0]
    
    def get_last_line_value(self, row_number):
        value = ""
        for cell in list(self.sheet.rows)[row_number - 1][:-2]:  # Exclude the last column
            if (cell.value
                    and str(cell.value).strip()
                    and re.fullmatch(r"[A-Z0-9]{7}", str(cell.value).strip())):
                value = cell.value
        return value

    def get_description(self, code):
        description = []
        for cell in list(self.sheet_extraction.columns)[1]:
            if (cell.value
                    and str(cell.value).strip()
                    and code in str(cell.value).strip()):
                description.append(cell.offset(column=8).value)
        return description


    def mise_en_forme(self):
        # 1. Define the border style once (efficient)
        thin_side = Side(style='thin')
        grid_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        
        # 2. Loop through columns from index 7 to the maximum column
        # +1 ensures the last column is included in range()
        for col_idx in range(1, self.sheet.max_column + 1):
            
            # 3. Loop through your target rows (rows 10, 11, and 12)
            # Use range(10, 13) to target rows 10, 11, and 12
            for row_idx in range(7, self.sheet.max_row + 1):
                
                # 4. Apply the border using the numeric coordinates
                self.sheet.cell(row=row_idx, column=col_idx).border = grid_border



    # ------------------------------------------------------------------
    # Main process
    # ------------------------------------------------------------------

    def start(self, output_path=None):
        results = []
        # self.load_excel()
        rows_to_delete = self.determine_rows_to_remove()

        if not rows_to_delete:
            save_path = output_path or self.excel_file_name
            if output_path:
                self.wb.save(save_path)
            return results, save_path

        self.delete_rows(rows_to_delete)

        for cell in self.create_col_before_last_column():

            value = self.get_last_line_value(cell.row)
            if value:

                cell.value = value
                description = self.get_description(value)
                desc_cell=cell.offset(column=2)
                if description and len(description) > 0:
                    desc_cell.value = ("; ".join(str(d) for d in description if d is not None))
                # results.append({
                #     "row": cell.row,
                #     "value": str(value),
                #     "status": "ok",
                #     "message": "Value extracted"
                # })
        self.delete_col()
        self.mise_en_forme()
        

        save_path = output_path or self.excel_file_name
        self.wb.save(save_path)
        return results, save_path


