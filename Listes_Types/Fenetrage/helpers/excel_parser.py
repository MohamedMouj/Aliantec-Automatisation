import openpyxl as op
import re

class excel_parser():
    def __init__(self, excel_file_name, context, parse_right=False):
        self.excel_file_name = excel_file_name
        self.context = context
        self.wb = None
        self.parse_right = parse_right

    def load_excel(self):
        self.wb = op.load_workbook(self.excel_file_name, data_only=True)
        sheet_to_remove=['NOTICE UTILISATION PTA PLM', 'ANNEXE 1', 'USER MANUEL PTA PLM', 'ANNEX 1(ENGLISH)', "NOTICE D'UTILISATION HNCT", 'SDP (LOGICALDIAGRAM)', 'HARNESS GROUP', 'FSC PHEV', 'COFORS', 'GRILLE APQP + PCP', 'DICO']
        
        for sheet in list(self.wb.worksheets):
            if sheet.title.strip().upper() in sheet_to_remove:
                self.wb.remove(sheet)

    def close(self):
        if self.wb:
            self.wb.close()

    def build_index(self):
        for sheet in self.wb.worksheets:
            for row in sheet.iter_rows(max_col=26):
                ref_cell, ref_string = self.find_first_valid_ref(row)
                if ref_cell is None and not ref_cell:
                    continue
                
                if ref_string in self.context.all_xml_references:
                    if ref_string not in self.context.excel_index:
                        self.context.excel_index[ref_string] = ref_cell
                else:
                    found = False
                    refs = []
                    while not found:
                        refs.append(ref_string)
                        ref_cell, ref_string = self.find_first_valid_ref(row, refs) 
                        if ref_cell is None:
                            break
                        if ref_string and ref_string in self.context.all_xml_references:
                            self.context.excel_index[ref_string] = ref_cell
                            found = True

    def find_first_valid_ref(self, row, jump_values=None):
        cells = reversed(row) if self.parse_right else row
        for cell in cells:
            cur = self.extract_reference_from_cell(cell)
            if jump_values and cur in jump_values:
                continue
            if cur is not None:
                return cell, cur
        return None, None

    def search_by_ref(self, ref):
        ref = str(ref).strip()
        if self.context.excel_index and ref in self.context.excel_index:
            return self.context.excel_index[ref]
        return None

    def is_delete_rule_triggered(self, cell):
        sheet = cell.parent 
        row = sheet[cell.row]
        if self.is_deleted(row, cell):
            return True
        
        # if cell.column <= 2:
        #     return False
        
        # left_cell = row[cell.column - 2] # column is 1-indexed
        # left_cell_2 = row[cell.column - 3]
        
        # # Check if neighbors are empty
        # if (left_cell.value is not None and str(left_cell.value).strip() != "") and (left_cell_2.value is not None and str(left_cell_2.value).strip() != ""):
        #     return False
            
        # # Check adjacent rows for patterns
        # for row_offset in [-4, 4]:
        #     target_row_idx = cell.row + row_offset
        #     if target_row_idx < 1 or target_row_idx > sheet.max_row:
        #         continue             
        #     neighbor_cell = sheet.cell(row=target_row_idx, column=cell.column - row_offset)
        #     if self.extract_reference_from_cell(neighbor_cell):
        #         return True 
                    
        return False 

    def extract_reference_from_cell(self, cell):
        # if cell is None or cell.value is None:
        #     return None
        # value = str(cell.value).strip()
        # if re.fullmatch(r"\d{10}", value):
        #     return value
        # match = re.match(r"^(\d{10})[-_*/.\\\\s]\d{2}", value)
        # if match:
        #     return match.group(1)
        # return None
        if cell is None or cell.value is None:
            return None
        value = str(cell.value).strip()
        match = re.search(r"\d{10}", value)
        if match:
            return match.group(0)
        return None

    def check_neighbors_detailed(self, cell):
        if self.parse_right:
            return self.check_right_neighbors_4_cells_detailed(cell)
        else:
            return self.check_left_neighbors_4_cells_detailed(cell)

    def check_left_neighbors_4_cells_detailed(self, cell):
        best_candidate = {
            "new_reference_detected": None,
            "neighbor_source_cell": None,
            "neighbor_distance": None,
            "raw_neighbor_value": None
        }
        for i in range(1, 9):
            if cell.column <= i:
                continue
            left_cell = cell.offset(row=0, column=-i)
            reference = self.extract_reference_from_cell(left_cell)
            if reference is not None:
                best_candidate["new_reference_detected"] = reference
                best_candidate["neighbor_source_cell"] = left_cell.coordinate
                best_candidate["neighbor_distance"] = i
                best_candidate["raw_neighbor_value"] = str(left_cell.value).strip()
        return best_candidate

    def check_right_neighbors_4_cells_detailed(self, cell):
        best_candidate = {
            "new_reference_detected": None,
            "neighbor_source_cell": None,
            "neighbor_distance": None,
            "raw_neighbor_value": None
        }
        for i in range(1, 9):
            try:
                right_cell = cell.offset(row=0, column=i)
                reference = self.extract_reference_from_cell(right_cell)
                if reference is not None:
                    best_candidate["new_reference_detected"] = reference
                    best_candidate["neighbor_source_cell"] = right_cell.coordinate
                    best_candidate["neighbor_distance"] = i
                    best_candidate["raw_neighbor_value"] = str(right_cell.value).strip()
            except:
                break
        return best_candidate

    def row_contains_red_cell(self, row, max_col=None):
        c=0
        for cell in row:
            if max_col and cell.column > max_col:
                break
            fill = cell.fill
            if str(cell.value).lower() in ["carryover", "traite", "", "treated"] and not (hasattr(cell.font, 'strike') and cell.font.strike):
                break
            if not fill:
                continue
                
            if hasattr(fill, 'start_color') and fill.start_color:
                color = str(fill.start_color.rgb).upper()
                if color in ["FFFF0000", "FF0000"] or fill.start_color.index == 2:
                    c+=1
                    if c==2:
                        return True
                    
        return False


        # for cell in row:
        #     fill = getattr(cell, "fill", None)

        #     if not fill:
        #         return False

        #     # Only real visible background fill
        #     if getattr(fill, "fill_type", None) != "solid":
        #         return False

        #     color = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)

        #     if not color:
        #         return False

        #     color_type = getattr(color, "type", None)

        #     # Case 1: RGB / ARGB color
        #     if color_type == "rgb":
        #         rgb = str(color.rgb).upper().replace("#", "")

        #         # openpyxl often gives ARGB like FFFF0000
        #         if len(rgb) == 8:
        #             rgb = rgb[-6:]

        #         if len(rgb) != 6:
        #             return False

        #         try:
        #             r = int(rgb[0:2], 16)
        #             g = int(rgb[2:4], 16)
        #             b = int(rgb[4:6], 16)
        #         except ValueError:
        #             return False

        #         # Red or red-like color
        #         return r >= 120 and r > g + 60 and r > b + 60

        #     # Case 2: Indexed color
        #     if color_type == "indexed":
        #         try:
        #             indexed = int(color.indexed)
        #         except (TypeError, ValueError):
        #             return False

        #         # Common Excel indexed reds
        #         return indexed in {2, 10}

        #     return False

    def row_contains_strike_cell(self, row, max_col=None):
        c=0
        for cell in row:
            if max_col and cell.column > max_col:
                break
            if cell.value ==None:
                continue
            font = cell.font
            if not font: continue
            if hasattr(font, 'strike') and font.strike:
                c+=1
                if c==2:
                    return True
        return False

    def row_contains_cancelled_status(self, row, max_col=None):
        keywords = {
            "delete",
            "deleted",
            "remove",
            "removed",
            "cancel",
            "cancelled",
            "canceled",
            "obsolete",
            "supprime",
            "supprimé",
            "a supprimer",
            "à supprimer",
            "annule",
            "annulé",
            "retire",
            "retiré",
        }
        for cell in row:
            if max_col and cell.column > max_col:
                break
            if cell.value: 
                val_up = str(cell.value).strip().lower()
                if any(kw in val_up for kw in keywords):
                    return True
        return False

    def is_deleted(self, row, cell):
        return self.row_contains_cancelled_status(row, max_col=cell.column) or \
               self.row_contains_red_cell(row, max_col=cell.column) or \
               self.row_contains_strike_cell(row, max_col=cell.column)
               #(self.row_contains_strike_cell(row, max_col=cell.column) and self.row_contains_red_cell(row, max_col=cell.column)) or \
