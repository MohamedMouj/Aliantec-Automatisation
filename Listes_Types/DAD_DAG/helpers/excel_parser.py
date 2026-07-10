

import openpyxl
from rapidfuzz import process, fuzz
import re
from collections import Counter
import math

class excel:
    def __init__(self, filename, context=None):
        self.filename = filename
        self.wb = None
        self.refs_desc_code = {}
        self.context = context

    def close(self):
        if self.wb:
            self.wb.close()

    def load(self):
        self.wb=openpyxl.load_workbook(self.filename, data_only=True)

    def detect_desc_col(self, ws):
        cols=[]
        for row in ws[1:50]:
            for cell in row[3:20]:
                if cell.value is not None and 20 <= len(str(cell.value)):
                    cols.append(cell.column)
                    break
        most_common_item = None
        if cols:
            most_common_item = Counter(cols).most_common(1)[0][0]
        return most_common_item # Returns 1

    def contains_code(self, value):
        if value is None:
            return False
        text = str(value).strip()
        
        
        pattern =  r'\b[A-Z0-9]{5}[&/+][A-Z0-9]{5}\b'  
       
        return bool(re.search(pattern, text, re.IGNORECASE | re.DOTALL))

    def detect_code_col(self, ws):
        cols=[]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and self.contains_code(str(cell.value)):
                    cols.append(cell.column)
                    break
            if len(cols)>=50:
                break
        most_common_item = None
        if cols:
            most_common_item = Counter(cols).most_common(1)[0][0]
        return most_common_item
    
    def build_refs_desc_mapping(self):
        fscf = self.context.fscfai_files.keys()
        
        for ws in self.wb:
            if ws.title in ["MINOR HARNESS", "Notice utilisation PTA PLM", "Annexe 1", "User manuel PTA PLM", "Annex 1(english)", "Notice d'utilisation HNCT", "SDP (LogicalDiagram)"]:
                continue
                
            col_number = self.detect_desc_col(ws)
            col_code = self.detect_code_col(ws)
            
            if not col_number:
                continue
         
            for row in ws.iter_rows():
                if col_number - 1 >= len(row):
                    continue
                    
                desc_cell = row[col_number - 1]
                code_cell = row[col_code - 1] if col_code else None
                
                if desc_cell.value is None:
                    continue
                
                has_forinfo = False
                for c in row:
                    if c.value:
                        val_str = str(c.value).lower()
                        if "forinfo" in val_str or "cancelled" in val_str:
                            has_forinfo = True
                            break
                if has_forinfo:
                    continue
                
                ref_cell = None
                for c in row:
                    if c.value is not None and re.match(r"\d{10}", str(c.value)) and (c.fill.start_color.rgb != "FF7DE1F5" or str(c.value) in self.context.all_xml_references) and str(c.value) in fscf:
                        if ref_cell is None:
                            ref_cell = c
                            break
                        
                            
                if not ref_cell:
                    continue
                    
                if code_cell is None:
                    code_cell = self.context.fscfai_files[str(ref_cell.value)]

                if getattr(ref_cell, 'font', None) and getattr(ref_cell.font, 'strikethrough', False):
                    continue
                    
                # if str(ref_cell.value) not in fscf and fallback_ref_cell:
                #     ref_cell = fallback_ref_cell
                
                if ref_cell and ref_cell.value is not None:
                    ref_key = str(ref_cell.value)
                    code_val = code_cell.value if hasattr(code_cell, "value") else code_cell
                    
                    self.refs_desc_code[ref_key] = str(code_val).strip()+str(desc_cell.value).strip()

    def search_ref(self, ref):
        for r in self.refs_desc_code.keys():
            if ref==r:
                return r, self.refs_desc_code[r]
        return None

    def get_ref_name(self, ref):
        return self.refs_desc.get(ref.value) if ref else None

    def contains_forinfo(self, cell):
        if cell and cell.parent:
            ws = cell.parent
            row_idx = cell.row
            
            # ws[row_idx] accurately loops through all columns for this specific row index
            for c in ws[row_idx]:
                if c.value:
                    cell_str = str(c.value).lower()
                    if "forinfo" in cell_str or "cancelled" in cell_str:
                        return True
        return False
            
    def get_code_by_ref():
        pass

    def find_matcheds(self, ref, desc, threshold=40):
        if 'DAG' in desc.upper():
            candidates = [
                d.upper() for r, d in self.refs_desc_code.items() if r != ref and 'DAD' in d
            ]
        else:
            candidates = [
                d.upper() for r, d in self.refs_desc_code.items() if r != ref and 'DAG' not in d and 'DAD' not in d and "LHD" not in d
            ]
        if not candidates:
            return None

        desc_upper = desc.upper()

        query=desc_upper
        if 'DAG' in desc_upper:
            candidates  = [c for c in candidates if 'DAD' in c.upper()]
            query = re.sub(r'\bDAG\b', 'DAD', desc, flags=re.IGNORECASE)
           

        if not candidates:
            return None

        scored=[]
        if "DAD" not in query.upper():
            scored = process.extract(query.upper(), candidates, scorer=fuzz.partial_ratio, limit=10)
        else:
            scored = process.extract(query.upper(), candidates, scorer=fuzz.ratio, limit=10)

        refs_desc={}
        if not (math.isclose(scored[0][1], scored[1][1], abs_tol=10)):
            refs_desc[self.get_ref_by_desc(scored[0][0])] = scored[0][0]
        else:
            for i in scored:
                refs_desc[self.get_ref_by_desc(i[0])] = i[0]
    
        return refs_desc if len(refs_desc)>0 else None

    def get_ref_by_desc_cell(self, desc_cell, jump_value=None):
        # FIX: Only search within the same row as the description cell
        for row in desc_cell.parent.iter_rows(min_row=desc_cell.row, max_row=desc_cell.row):
            for cell in row:
                if jump_value and str(cell.value)==jump_value:
                    continue
                if cell.value is not None and re.match(r"\d{10}", str(cell.value)):
                    return cell
        return None
   
    def get_ref_by_desc(self, desc):
        for r, d in self.refs_desc_code.items():
            if desc.upper()==d.upper():
                return r
            
    def get_fus_name(self, desc):
        if desc:
            if desc[:2].upper()=='PR': return desc=='PR'
            if desc[:4].upper()=='EHAB': return desc[:5].upper()
            return None
        
    def start(self, ref, old_xml_path=None):
        ref_desc = self.search_ref(ref)
        if ref_desc\
        and ((('DAG' in ref_desc[1].upper()) and not('DAD' in ref_desc[1].upper())) \
        or ('50-PB' in old_xml_path or '60-PA' in old_xml_path or '62-PRG' in old_xml_path or '67-PRD' in old_xml_path) \
        or ('LHD' in ref_desc[1].upper())):
            refs = self.find_matcheds(ref_desc[0], ref_desc[1]) if ref_desc else None
            return refs
        return None

    