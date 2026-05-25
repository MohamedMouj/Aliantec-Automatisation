import openpyxl
from rapidfuzz import process
import re

class excel:
    def _init_(self, filename):
        self.filename=filename
    
    def load(self):
        self.wb=openpyxl.load_workbook(self.filename)

    def search_ref(self, ref):
        for ws in self.wb:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == ref:
                        return row

    def get_ref_name(self, row):
        if row:
            for cell in row:
                if cell.value is not None and 30 <= len(str(cell.value))<=70:
                    return cell


    def find_target_cell_column(self, cell):
        col_idx = cell.column
        list_desc=[]
        map_cell_desc ={}
        for cell in col_idx:
            next_cell = self.wb.cell(row=cell.row+1, column=col_idx)
            if next_cell.value is None:
                map_cell_desc[next_cell]=cell.value
                list_desc.append(next_cell.value)
        best_match = process.extractOne(cell.value, list_desc, scorer=fuzz.WRatio)
        return map_cell_desc[best_match[0]]

    def get_ref_by_desc(self,desc_cell):
        for cell in desc_cell.row:
            if re.match(r"[\d{10}]", cell.value):
                return cell
        
    def start(self, ref, filename):
        self.filename=filename
        self.load()
        ref_cell = self.search_ref(ref)
        desc_cell = self.get_ref_name(ref_cell)
        target_cell = self.find_target_cell_column(self, desc_cell)
        target_ref = self.get_ref_by_desc(target_cell)
        return target_ref

excel1= excel()
print(excel1.start("9872956580", "C:/Users/User/OneDrive/Bureau/OV512  DAG DAD/OV512  DAG DAD/PTA_OV512E MCA_F0226.xlsm"))

    
    


    
            
        